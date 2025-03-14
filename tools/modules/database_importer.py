import openpyxl
import json
from datetime import datetime
from urllib.parse import urlparse


def log(message):
    timestamp = datetime.now().strftime("[%Y-%m-%d %H:%M:%S]")
    print(f"{timestamp} {message}")


# Open the Excel file
log("Loading Excel file...")
workbook = openpyxl.load_workbook("../ExporterSheet.xlsx")
worksheet = workbook.active

# Load existing data from the output JSON file if it exists
log("Loading JSON data...")
try:
    with open("../Compromised-Discord-Accounts.json", "r", encoding="utf-8") as file:
        data = json.load(file)
except FileNotFoundError:
    data = {}

# Create a set of existing DISCORD_IDs from the loaded JSON data
existing_discord_ids = {account["DISCORD_ID"] for account in data.values()}

# Get the number of existing cases in JSON and Excel
json_case_count = len(data)
excel_case_count = sum(1 for _ in worksheet.iter_rows(min_row=2, values_only=True))
log(f"Found {json_case_count} cases in JSON file.")
log(f"Found {excel_case_count} cases in Excel sheet.")

# Iterate over the rows in the worksheet
new_cases = 0
for row_number, row in enumerate(
    worksheet.iter_rows(min_row=2, values_only=True), start=1
):
    (
        NOUMBER,
        FOUND_ON,
        DISCORD_ID,
        USERNAME,
        BEHAVIOUR,
        TYPE,
        METHOD,
        TARGET,
        PLATFORM,
        SURFACE_URL,
        REGION,
        STATUS,
    ) = row

    discord_id_str = str(DISCORD_ID) if DISCORD_ID is not None else "Unknown"
    if discord_id_str in existing_discord_ids or discord_id_str == "Unknown":
        continue

    if not any(
        [
            FOUND_ON,
            DISCORD_ID,
            USERNAME,
            BEHAVIOUR,
            TYPE,
            METHOD,
            TARGET,
            PLATFORM,
            SURFACE_URL,
            REGION,
            STATUS,
        ]
    ):
        continue

    new_cases += 1
    found_on_str = FOUND_ON.strftime("%Y-%m-%d") if FOUND_ON else "Unknown"
    surface_url_domain = urlparse(SURFACE_URL).netloc if SURFACE_URL else ""
    non_ascii_username = not USERNAME.isascii() if USERNAME else False

    account = {
        "CASE_NUMBER": f"{row_number}",
        "FOUND_ON": found_on_str,
        "DISCORD_ID": discord_id_str,
        "USERNAME": USERNAME if USERNAME is not None else "Unknown",
        "ACCOUNT_STATUS": "",
        "BEHAVIOUR": BEHAVIOUR if BEHAVIOUR is not None else "Unknown",
        "ATTACK_METHOD": TYPE if TYPE is not None else "Unknown",
        "ATTACK_VECTOR": METHOD if METHOD is not None else "Unknown",
        "ATTACK_GOAL": TARGET if TARGET is not None else "Unknown",
        "ATTACK_SURFACE": PLATFORM if PLATFORM is not None else "Unknown",
        "SUSPECTED_REGION_OF_ORIGIN": REGION if REGION is not None else "Unknown",
        "SURFACE_URL": SURFACE_URL if SURFACE_URL is not None else "Unknown",
        "SURFACE_URL_DOMAIN": surface_url_domain,
        "SURFACE_URL_STATUS": STATUS if STATUS is not None else "Unknown",
        "FINAL_URL": "",
        "FINAL_URL_DOMAIN": "",
        "FINAL_URL_STATUS": "",
        "NON_ASCII_USERNAME": non_ascii_username,
        "LAST_CHECK": "",
    }

    data[f"ACCOUNT_NUMBER_{len(data) + 1}"] = account
    existing_discord_ids.add(discord_id_str)

log(f"{new_cases} new cases will be added to JSON file.")

# Write the updated JSON data to a file
log("Saving updated JSON data...")
with open("../Compromised-Discord-Accounts.json", "w", encoding="utf-8") as file:
    json.dump(data, file, indent=4, ensure_ascii=False)
log("Update complete.")
